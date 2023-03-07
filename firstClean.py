from openpyxl import load_workbook, Workbook
from difflib import SequenceMatcher
import pandas
from functions import *

"""NOTE:T_ indicates variable is from Template document
        S_ indicates variale is from Source document
        
        """

# Load all necessary workbooks
siteName = "AucklandWarMemorial" # "HealthcareHB"
T_Workbook = load_workbook(filename= siteName + "/SMS-Asset-Collection-NZ20-" + siteName + ".xlsx")
S_Workbook = load_workbook(filename= siteName + "/SMS-Assets-Raw-" + siteName + ".xlsx")
N_Workbook = Workbook()
N_Sheet = N_Workbook.active

# Pull relevant Sheets
T_Listing = T_Workbook['Listing']
S_assets = S_Workbook['APAC - Assets by Location']

# Relvant Columns in SMSassets
S_catalogFamily = S_assets['G']
S_productName = S_assets['H']
S_equipmentTag = S_assets['N']
S_egressFactor = S_assets['L']
S_quantity = S_assets['F']
S_description = S_assets['J']

# print(S_productName[0])
# print(S_productName[1])
# print(S_productName[2])

# GAR categories (Done off the 'Name' columns not 'GAR-Product') category
T_controlsGARFam = T_Listing['A'][1:42]
T_forgeGARFam = T_Listing['D'][1:46]
T_ICTGARFam = T_Listing['G'][1:15]
T_securityGARFam = T_Listing['J'][1:22]
T_fireGARFam = T_Listing['M'][1:27]

# Dictionary containing all the types of GAR fams for each of the families
T_assetFamilies = {'Controls':T_controlsGARFam , 'Forge':T_forgeGARFam, 'ICT':T_ICTGARFam, 'Security':T_securityGARFam, 'Fire Alarm':T_fireGARFam}

# for fammy in assetFamilies:
#     print(fammy)
#     for GARname in assetFamilies[fammy]:
#         print(GARname.value)

# Set start and end of total SMS database
start = 1
print(len(S_catalogFamily))
end = len(S_catalogFamily)

exportTitles = [['Index', 'Tag/ID', 'Quantity', 'Description', 'Asset Swap', 'Similarity', 'Product Name Swap', 'Similarity']]
exportList = [['Index', 'Tag/ID', 'Quantity', 'Description', 'Asset Swap', 'Similarity', 'Product Name Swap', 'Similarity']]
prodSimDict = {}

# Iterate through the whole database
for i in range(start, end):
    family = S_catalogFamily[i].value
    prodName = S_productName[i].value
    # print('Family: ', family, '\nprodName: ', prodName)
    
    fam, famRank = familyMatch(family, T_assetFamilies)
    familyEntry = '"{source}" -> "{template}"'.format(source=family, template=fam)

    prod, prodRank = productMatch(prodName, T_assetFamilies[fam])
    productEntry = '"{source}" -> "{template}"'.format(source=prodName, template=prod)
    
    item = [i, S_equipmentTag[i].value, S_quantity[i].value, S_description[i].value, familyEntry, famRank, productEntry, float(prodRank)]
    
    exportList.append(item)

print(len(exportList))
print(len(exportTitles))
sortedExportList = sorted(exportList[1:], key=lambda row: (row[5], row[7]))
# Add title row
sortedExportList = exportTitles + sortedExportList
print(len(sortedExportList))
FullDataFrame = pandas.DataFrame(sortedExportList)
FullDataFrame.to_excel(excel_writer = siteName + "/sortedList-" + siteName + ".xlsx")

            
            

            

